import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

#dictionary created for suppliers and products
products_per_supplier = {}
#print(product_list.max_row)

#range default starts from 0, last is exclusive
for product_row in range(2, product_list.max_row+1):
    supplier_name = product_list.cell(product_row, 4).value

    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print("adding a mew supplier")
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)
